import pandas as pd
import os
import time
from datetime import datetime
import openpyxl
import json
import tldextract
import requests
from urllib.parse import urlparse
from domains import ias_domains, dv_domains, moat_domains, human_security_domains, pixalate_domains
MAX_CHAR_LIMIT = 32767
MAX_URL_LIMIT = 2079
PHASE = "Controlled"




def checkPayloads(brand, phase, website, request_payload, response_payload):
    payload_directory = './payloads'
    if not os.path.exists(payload_directory):
        os.makedirs(payload_directory)
    payload_directory = os.path.join(os.getcwd(), 'payloads')


    try:
        if len(request_payload) > MAX_CHAR_LIMIT:
            timestamp = datetime.now().strftime("%H%M%S%f")[:-3]
            filename = f"{brand}_{phase}_{website}_request_{timestamp}.txt"
            file_path = os.path.join(payload_directory, filename)
            # print("YOO check this out in request\n", request_payload, "\n\n")
            with open(file_path, 'w') as file:
                file.write(request_payload)
                file.flush()
            time.sleep(0.5)
            new_request_payload = file_path
        else:
            new_request_payload = request_payload
    except Exception as e:
        print("ERROR"*100)
        print("ERROR IN REQUEST BLOCK!!!!!!!!!")
        print("ERROR", e, "\n")


    try:
        if len(response_payload) > MAX_CHAR_LIMIT:
            timestamp = datetime.now().strftime("%H%M%S%f")[:-3]
            filename = f"{brand}_{phase}_{website}_response_{timestamp}.txt"
            file_path = os.path.join(payload_directory, filename)
            # print("YOO check this out in response\n", response_payload, "\n\n")
            with open(file_path, 'w') as file:
                file.write(response_payload)
                file.flush()
            time.sleep(0.5)
            new_response_payload = file_path
        else:
            new_response_payload = response_payload
    except Exception as e:
        print("ERROR"*100)
        print("ERROR IN REQUEST BLOCK!!!!!!!!!")
        print("ERROR", e, "\n")


    return new_request_payload, new_response_payload





def storeDataToExcel(brand, phase, website, har_path, excel_path_master, excel_path_data, entry, request_url, domainURL, domain, brand_safety_service, brand_safety_idx):
    
    # Extract payload or additional data if available
    response = entry.get('response', {})
    redirects = response.get('redirectURL', '')
    request_payload = entry.get('request', {}).get('postData', {}).get('text', '')
    response_payload = response.get('content', {}).get('text', '')
    status_code = response.get('status', '')


    # MASTER EXCEL -- UPDATE!
    df_master = pd.read_excel(excel_path_master)
    # Update corresponding domain count
    website_row = df_master[df_master['Website'] == website].index[0]
    df_master.at[website_row, domain] = df_master.at[website_row, domain] + 1
    # Update corresponding brand safety service count
    rr_mapping = {
        1: "IAS Domains R&R",
        2: "DoubleVerify Domains R&R",
        3: "Moat Domains R&R",
        4: "Human Security Domains R&R",
        5: "Pixalate Domains R&R",
    }
    rr_column = rr_mapping[brand_safety_idx]
    df_master.at[website_row, rr_column] += 1
    df_master.to_excel(excel_path_master, index=False)



    # DATA EXCEL -- UPDATE!
    request_url = request_url[:MAX_URL_LIMIT]
    redirects = redirects[:MAX_URL_LIMIT]
    new_request_payload, new_response_payload = checkPayloads(brand, phase, website, request_payload, response_payload)

    result = {
        "Brand": brand,
        "Phase": phase, 
        "Website": website, 
        "Request URL": request_url, 
        "Domain": domainURL, 
        "Brand Safety Domain": domain, 
        "Brand Safety Service": brand_safety_service,
        "Response Status": status_code, 
        "Redirects": redirects, 
        "Request Payload": new_request_payload, 
        "Response Payload": new_response_payload
    }
    new_entry = pd.DataFrame([result])
    df_data = pd.read_excel(excel_path_data)
    df_data = pd.concat([df_data, new_entry], ignore_index=True)
    df_data.to_excel(excel_path_data, index=False)






def analyze_har(brand, phase, website, har_path, excel_path_master, excel_path_data):

    # Load domain data for IAS, DoubleVerify, and Moat
    # IMPORTED!


    # Load HAR file data
    try:
        with open(har_path, 'r') as f:
            har_data = json.load(f)
    except Exception as e:
        return 0
    
    # Extract entries!!
    entries = har_data.get('log', {}).get('entries', [])



    # Analyze all requests and responses/redirects in the HAR file
    for entry in entries:

        request_url = entry.get('request', {}).get('url', '')


        # Extract the domain from the request URL
        parsed_url = urlparse(request_url)
        domainURL = parsed_url.netloc

        

        ### INTEGRAL AD SCIENCE
        for domain in ias_domains:
            if domainURL.endswith(domain):
                storeDataToExcel(brand, phase, website, har_path, excel_path_master, excel_path_data, entry, request_url, domainURL, domain, "IAS", 1)
        
        ### DOUBLE VERIFY
        for domain in dv_domains:
            if domainURL.endswith(domain):
                storeDataToExcel(brand, phase, website, har_path, excel_path_master, excel_path_data, entry, request_url, domainURL, domain, "DoubleVerify", 2)

        ### MOAT
        for domain in moat_domains:
            if domainURL.endswith(domain):
                storeDataToExcel(brand, phase, website, har_path, excel_path_master, excel_path_data, entry, request_url, domainURL, domain, "Moat", 3)

        ### HUMAN SECURITY
        for domain in human_security_domains:
            if domainURL.endswith(domain):
                storeDataToExcel(brand, phase, website, har_path, excel_path_master, excel_path_data, entry, request_url, domainURL, domain, "Human Security", 4)

        ### PIXALATE
        for domain in pixalate_domains:
            if domainURL.endswith(domain):
                storeDataToExcel(brand, phase, website, har_path, excel_path_master, excel_path_data, entry, request_url, domainURL, domain, "Pixalate", 5)


    return 1





def iterateBrandFolder(BRAND):

    # Path to Excels!!!
    excel_path_master = BRAND.lower() + "_" + PHASE.lower() + "_" + "master.xlsx"
    excel_path_master = os.path.join(master_combined_directory, excel_path_master)
    excel_path_data = BRAND.lower() + "_" + PHASE.lower() + "_" + "data.xlsx"
    excel_path_data = os.path.join(data_combined_directory, excel_path_data)



    # Create Master Excel
    core_columns = ["Brand", "Phase", "Website", "HAR File Present"]
    rr_columns = [
        "IAS Domains R&R", 
        "DoubleVerify Domains R&R", 
        "Moat Domains R&R", 
        "Human Security Domains R&R", 
        "Pixalate Domains R&R"
    ]
    domain_columns = ias_domains + dv_domains + moat_domains + human_security_domains + pixalate_domains   # 48 in number!
    columnsExcel = core_columns + rr_columns + domain_columns
    df = pd.DataFrame(columns = columnsExcel)
    df.to_excel(excel_path_master, index=False)

    # Create Data Excel
    columnsExcel = ["Brand", "Phase", "Website", "Request URL", "Domain", "Brand Safety Domain", "Brand Safety Service", "Response Status", "Redirects", "Request Payload", "Response Payload"]
    df = pd.DataFrame(columns = columnsExcel)
    df.to_excel(excel_path_data, index=False)





    # Path to the directory containing the website folders (IMPORTANT!!!!)
    ########################## UPDATABLE #########################################
    ######################## ERROR PRONE??????S ###########################################
    parent_dir = os.path.join(os.getcwd(), BRAND)


    # ITERATE OVER /BRAND/PHASE
    for folder_name in os.listdir(parent_dir):

        if folder_name == "nan":
            continue

        # OPEN MASTER EXCEL (check if the excel file exists)
        if os.path.exists(excel_path_master):
            workbookMaster = openpyxl.load_workbook(excel_path_master)
            sheetMaster = workbookMaster.active
            
            # APPEND A NEW ROW WITH 0 VALUES!!!!
            BPW = [BRAND, PHASE, folder_name, 1]
            rr_columns = [0] * 5  # Corresponding to R&R columns - 5 Brand Safety Service
            domain_columns = [0] * 48  # For all domain-specific columns
            new_row = BPW + rr_columns + domain_columns
            sheetMaster.append(new_row)
            workbookMaster.save(excel_path_master)

        else:
            raise FileNotFoundError(f"Error: The file '{excel_path_master}' does not exist. This should not have happened by the way?!!")



        # NAVIGATE TO HAR FILE
        folder_path = os.path.join(parent_dir, folder_name)
        file_name = BRAND.lower() + "_" + folder_name + "_har.json"
        har_path = os.path.join(folder_path, file_name)


        # CHECK IF THE HAR FILE WAS COLLECTED
        if os.path.exists(har_path):
            print(f"HAR File DOES exist: {har_path}\n")
            flag = analyze_har(BRAND, PHASE, folder_name, har_path, excel_path_master, excel_path_data)
            if flag == 0:
                last_row = sheetMaster.max_row
                sheetMaster.cell(row=last_row, column=4, value=0)
                workbookMaster.save(excel_path_master)
                print("CHARMAP ERROR................................\nUnicode blah blah blah error!!!\n")


        else:
            print(f"HAR File does not exist: {har_path}\n")
            last_row = sheetMaster.max_row
            sheetMaster.cell(row=last_row, column=4, value=0)
            workbookMaster.save(excel_path_master)







# ---------------------------------------------- MAIN -----------------------------------------------------------------------------

master_combined_directory = './Master Excel Combined'
if not os.path.exists(master_combined_directory):
    os.makedirs(master_combined_directory)
master_combined_directory = os.path.join(os.getcwd(), 'Master Excel Combined')


data_combined_directory = './Data Excel Combined'
if not os.path.exists(data_combined_directory):
    os.makedirs(data_combined_directory)
data_combined_directory = os.path.join(os.getcwd(), 'Data Excel Combined')



# Scalable knobs... (add a iterator)
for brand in os.listdir(os.getcwd()):
    if brand == "analysisHAR.py":
        continue
    if brand == "Master Excel Combined":
        continue
    if brand == "Data Excel Combined":
        continue
    if brand == "payloads":
        continue
    if brand == "domains.py":
        continue
    if brand == "__pycache__":    
        continue
    # print(brand)
    iterateBrandFolder(brand)